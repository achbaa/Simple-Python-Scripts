import os
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import asyncio
from openai import AsyncOpenAI
import time
from pathlib import Path
import argparse
import re

async def categorize_batch(client, batch_data, system_prompt, user_prompt_template, model="gpt-4o-mini"):
    """Process a batch of items asynchronously using the OpenAI API"""
    try:
        # Create a numbered list of items for clear identification
        numbered_items = []
        for i, item in enumerate(batch_data, 1):
            numbered_items.append(f"ITEM {i}: {item}")
        
        # Join all items with clear separators
        batch_text = "\n\n".join(numbered_items)
        
        # Make a single API call for the batch with optimized prompt
        response = await client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Categorize each of the following items. For each item, respond with ONLY the category name - no numbering, no punctuation, no explanations:\n\n{batch_text}"}
            ],
            max_tokens=100 * len(batch_data),
            temperature=0.3  # Lower temperature for more consistent results
        )
        
        # Extract categories from response
        response_text = response.choices[0].message.content.strip()
        
        # Try to extract categories using the ITEM pattern first
        pattern = r"ITEM\s+(\d+):\s*(.*?)(?=\s*ITEM\s+\d+:|$)"
        matches = re.findall(pattern, response_text, re.DOTALL)
        
        if matches and len(matches) == len(batch_data):
            # Sort by item number and extract just the categories
            sorted_matches = sorted(matches, key=lambda x: int(x[0]))
            categories = [match[1].strip() for match in sorted_matches]
        else:
            # Fallback: split by lines and clean up
            categories = response_text.split('\n')
            categories = [c for c in categories if c.strip()]
            
            # Clean up any item numbers or prefixes
            cleaned_categories = []
            for category in categories:
                # Remove item numbers if present
                category = re.sub(r'^(ITEM\s+\d+:|\d+[.)])\s*', '', category).strip()
                # Remove any quotes, periods at the end, etc.
                category = category.strip('"\'.,;:').strip()
                cleaned_categories.append(category)
            
            categories = cleaned_categories
        
        # If we got fewer categories than items, pad with empty strings
        while len(categories) < len(batch_data):
            categories.append("")
            
        return categories
        
    except Exception as e:
        print(f"Error processing batch: {str(e)}")
        return ["Error: " + str(e)] * len(batch_data)

async def process_batches(client, data, batch_size, system_prompt, user_prompt_template, model, max_concurrent=4):
    """Process multiple batches concurrently with rate limiting"""
    semaphore = asyncio.Semaphore(max_concurrent)
    
    async def process_with_semaphore(batch):
        async with semaphore:
            return await categorize_batch(client, batch, system_prompt, user_prompt_template, model)
    
    # Split data into batches
    batches = [data[i:i + batch_size] for i in range(0, len(data), batch_size)]
    
    # Process all batches concurrently
    tasks = [process_with_semaphore(batch) for batch in batches]
    results = await asyncio.gather(*tasks)
    
    # Flatten results
    return [cat for batch_result in results for cat in batch_result]

async def async_main():
    """Async main function to run the script."""
    # Set up argument parser for quick mode
    parser = argparse.ArgumentParser(description="Excel Data Categorization Tool")
    parser.add_argument("--quick", action="store_true", help="Run in quick mode with default settings")
    parser.add_argument("--file", type=str, help="Path to Excel file")
    parser.add_argument("--source", type=int, help="Source column number")
    parser.add_argument("--target", type=int, help="Target column number")
    parser.add_argument("--start", type=int, help="Starting row number")
    parser.add_argument("--batch", type=int, default=5, help="Batch size for processing (default: 5)")
    parser.add_argument("--model", type=str, default="gpt-4o-mini", help="OpenAI model to use (default: gpt-4o-mini)")
    parser.add_argument("--categories", type=str, help="Comma-separated list of allowed categories")
    args = parser.parse_args()

    print("Excel Data Categorization Tool")
    print("------------------------------")
    
    # Load API key from .env file
    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        print("Error: OPENAI_API_KEY not found in .env file.")
        return
    
    # Initialize OpenAI client
    client = AsyncOpenAI(api_key=api_key)
    
    # Get file path
    file_path = args.file if args.file else input("Enter the path to your Excel file: ")
    
    try:
        # Load the workbook and get the active sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Quick mode skips the preview and uses defaults
        if not args.quick:
            # Ask which row contains headers
            header_row = 1
            try:
                print("\nPreview of first few rows:")
                for r in range(1, min(6, sheet.max_row + 1)):
                    row_preview = []
                    for c in range(1, min(8, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=r, column=c).value
                        row_preview.append(str(cell_value) if cell_value is not None else "None")
                    print(f"Row {r}: {' | '.join(row_preview)}")
                    
                header_row_input = input("\nWhich row contains the column headers? (default: 1): ")
                if header_row_input.strip():
                    header_row = int(header_row_input)
            except ValueError:
                print("Invalid input. Using row 1 as headers.")
        else:
            header_row = 1
        
        # Get column headers for reference
        headers = []
        for cell in sheet[header_row]:
            header_value = cell.value if cell.value is not None else f"Column {openpyxl.utils.get_column_letter(cell.column)}"
            headers.append(header_value)
        
        if not args.quick:
            print("\nAvailable columns:")
            for i, header in enumerate(headers, 1):
                print(f"{i}. {header}")
        
        # Get source column
        if args.source:
            source_col_idx = args.source
        else:
            while True:
                try:
                    source_col_idx = int(input("\nEnter the number of the column to categorize: "))
                    if 1 <= source_col_idx <= len(headers):
                        break
                    print(f"Please enter a number between 1 and {len(headers)}.")
                except ValueError:
                    print("Please enter a valid number.")
        source_col = openpyxl.utils.get_column_letter(source_col_idx)
        
        # Get starting row
        if args.start:
            start_row = args.start
        else:
            while True:
                try:
                    start_row = int(input(f"Enter the row number to start categorizing from (after header row {header_row}): "))
                    if start_row > header_row:
                        break
                    print(f"Please enter a number greater than the header row ({header_row}).")
                except ValueError:
                    print("Please enter a valid number.")
        
        # Get target column
        if args.target:
            target_col_idx = args.target
        else:
            while True:
                try:
                    target_col_idx = int(input("Enter the number of the column to put categorizations: "))
                    if 1 <= target_col_idx <= len(headers):
                        break
                    print(f"Please enter a number between 1 and {len(headers)}.")
                except ValueError:
                    print("Please enter a valid number.")
        target_col = openpyxl.utils.get_column_letter(target_col_idx)
        
        # Set batch size
        batch_size = args.batch
        
        # Set model
        model = args.model
        
        # Get allowed categories if specified
        allowed_categories = None
        if args.categories:
            allowed_categories = [cat.strip() for cat in args.categories.split(',')]
        elif not args.quick:
            use_allowed_categories = input("Would you like to specify allowed categories? (y/n): ").lower() == 'y'
            if use_allowed_categories:
                categories_input = input("Enter comma-separated list of allowed categories: ")
                allowed_categories = [cat.strip() for cat in categories_input.split(',')]
        
        # Build the optimized system prompt
        if allowed_categories:
            categories_list = "\n".join([f"- {cat}" for cat in allowed_categories])
            system_prompt = (
                "You are an expert data categorization system. Your task is to analyze each item and assign it to exactly one category "
                "from the following list of allowed categories:\n\n"
                f"{categories_list}\n\n"
                "IMPORTANT INSTRUCTIONS:\n"
                "1. Use ONLY categories from the provided list.\n"
                "2. Provide ONLY the category name - no explanations, no numbering, no punctuation.\n"
                "3. Each response should be the exact category name as listed above.\n"
                "4. For each item, respond with the category name only.\n"
                "5. If an item doesn't clearly fit any category, choose the closest match from the allowed categories."
            )
        else:
            system_prompt = (
                "You are an expert data categorization system. Your task is to analyze each item and assign the most "
                "appropriate category label.\n\n"
                "IMPORTANT INSTRUCTIONS:\n"
                "1. Provide ONLY the category name - no explanations, no numbering, no punctuation.\n"
                "2. Be consistent with category names across all items.\n"
                "3. Use short, clear category names.\n"
                "4. For each item, respond with the category name only.\n"
                "5. Do not add quotes, periods, or any other characters to the category names."
            )
        
        user_prompt_template = "{text}"
        
        # Ask for custom prompt if not in quick mode
        if not args.quick:
            use_custom_prompt = input("Would you like to use a custom prompt for categorization? (y/n): ").lower() == 'y'
            if use_custom_prompt:
                system_prompt = input("Enter system prompt (instructions for the AI): ") or system_prompt
                user_prompt_template = input("Enter user prompt template (use {text} as placeholder): ") or user_prompt_template
        
        # Ask if user wants to save to a new file if not in quick mode
        save_to_new_file = False
        output_file_path = file_path
        
        if not args.quick:
            save_to_new_file = input("Would you like to save results to a new file? (y/n): ").lower() == 'y'
            
        # If saving to a new file, determine the output path
        if save_to_new_file:
            if args.quick and args.file:
                # In quick mode with specified file, use default output location
                script_dir = Path(__file__).parent
                output_dir = script_dir / "Output files"
                output_dir.mkdir(exist_ok=True)
                file_name = Path(file_path).name
                default_new_path = output_dir / f"{Path(file_name).stem}_categorized{Path(file_name).suffix}"
                output_file_path = str(default_new_path)
            elif not args.quick:
                # In interactive mode, ask for output location with default in Output files folder
                script_dir = Path(__file__).parent
                output_dir = script_dir / "Output files"
                output_dir.mkdir(exist_ok=True)
                file_name = Path(file_path).name
                default_new_path = output_dir / f"{Path(file_name).stem}_categorized{Path(file_name).suffix}"
                output_file_path = input(f"Enter new file path (default: {default_new_path}): ") or str(default_new_path)
        else:
            # If not saving to a new file, still save to Output files folder by default
            script_dir = Path(__file__).parent
            output_dir = script_dir / "Output files"
            output_dir.mkdir(exist_ok=True)
            file_name = Path(file_path).name
            output_file_path = output_dir / f"{Path(file_name).stem}_categorized{Path(file_name).suffix}"
        
        # Count total rows to process
        max_row = sheet.max_row
        total_rows = max_row - start_row + 1
        
        # Preview what will be done
        print("\nOperation Summary:")
        print(f"- Source column: {headers[source_col_idx-1]} (Column {source_col})")
        print(f"- Target column: {headers[target_col_idx-1]} (Column {target_col})")
        print(f"- Starting row: {start_row}")
        print(f"- Total rows to process: {total_rows}")
        print(f"- Batch size: {batch_size}")
        print(f"- Model: {model}")
        if allowed_categories:
            print(f"- Allowed categories: {', '.join(allowed_categories)}")
        print(f"- Output file: {output_file_path}")
        print("\nNote: Cells with existing values in the target column will be skipped.")
        
        # Ask for confirmation if not in quick mode
        if not args.quick:
            confirm = input("\nProceed with categorization? (y/n): ").lower()
            if confirm != 'y':
                print("Operation cancelled.")
                return
        
        print(f"\nProcessing {total_rows} rows in batches of {batch_size}...")
        
        # Track progress
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        # Collect all data to process
        all_data = []
        row_map = []  # Keep track of which rows to update
        
        current_row = start_row
        while current_row <= max_row:
            cell_value = sheet[f"{source_col}{current_row}"].value
            target_cell = sheet[f"{target_col}{current_row}"]
            
            # Skip if target cell already has a value
            if target_cell.value:
                print(f"Row {current_row}/{max_row}: Skipped (target cell already has value)")
                skipped_count += 1
            elif cell_value:
                all_data.append(cell_value)
                row_map.append(current_row)
            else:
                print(f"Row {current_row}/{max_row}: Empty cell, skipping")
                skipped_count += 1
            
            current_row += 1
        
        if all_data:
            try:
                # Process all data concurrently in batches
                categories = await process_batches(
                    client,
                    all_data,
                    batch_size,
                    system_prompt,
                    user_prompt_template,
                    model
                )
                
                # Update cells with categories
                for i, row_num in enumerate(row_map):
                    if i < len(categories):
                        category = categories[i]
                        
                        # Validate against allowed categories if specified
                        if allowed_categories and category and category not in allowed_categories:
                            # Try to find the closest match
                            closest_match = min(allowed_categories, key=lambda x: abs(len(x) - len(category)))
                            print(f"Row {row_num}/{max_row}: Warning - '{category}' not in allowed categories, using '{closest_match}' instead")
                            category = closest_match
                        
                        sheet[f"{target_col}{row_num}"] = category
                        processed_count += 1
                        print(f"Row {row_num}/{max_row}: Categorized as '{category}'")
                    else:
                        error_count += 1
                        print(f"Row {row_num}/{max_row}: Error - no category returned")
                    
                    # Save periodically
                    if (i + 1) % (batch_size * 5) == 0:
                        workbook.save(output_file_path)
                        print(f"Progress saved at row {row_num}/{max_row}")
                
            except Exception as e:
                print(f"Error processing data: {str(e)}")
                error_count += len(all_data)
        
        # Final save
        workbook.save(output_file_path)
        
        # Print summary
        print("\nCategorization complete!")
        print(f"- Processed: {processed_count} rows")
        print(f"- Skipped: {skipped_count} rows (empty cells or existing values)")
        print(f"- Errors: {error_count} rows")
        print(f"- Excel file has been saved to: {output_file_path}")
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    asyncio.run(async_main())
